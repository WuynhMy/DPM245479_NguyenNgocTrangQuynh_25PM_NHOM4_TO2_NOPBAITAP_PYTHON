'''Xử lý đọc Excel File
Yêu cầu:
Sử dụng thư viện openpyxl để đọc file excel ở câu trước.'''
from openpyxl import load_workbook
wb = load_workbook('demo.xlsx')
Trang 67/84
print (wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.values:
 for value in row:
 print(value,"\t",end='')
 print("")