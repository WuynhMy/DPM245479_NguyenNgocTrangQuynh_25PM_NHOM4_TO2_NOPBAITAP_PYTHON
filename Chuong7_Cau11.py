'''Xử lý Excel File - Viết phần mềm Quản Lý Nhân Viên
Yêu cầu:
Viết phần mềm quản lý Nhân viên lưu bằng Excel. Mỗi nhân viên có Mã, Tên, Tuổi.
− Phần mềm cho phép lưu Nhân viên vào File Excel
− Phần mềm cho phép đọc danh sách Nhân viên trong File Excel
− Phần mềm cho phép sắp xếp Nhân viên theo Tuổi tăng dần
Cấu trúc của File Excel như sau:'''
import openpyxl
from openpyxl import Workbook
import os


# ==============================
# ĐỊNH NGHĨA LỚP DỮ LIỆU
# ==============================
class NhanVien:
    def __init__(self, ma, ten, tuoi):
        self.ma = ma
        self.ten = ten
        self.tuoi = int(tuoi)


# ==============================
# LỚP QUẢN LÝ NHÂN VIÊN
# ==============================
class QuanLyNhanVien:
    def __init__(self, filename="nhanvien.xlsx"):
        self.filename = filename
        self.ds_nv = []
        self.doc_file()

    # ---- Đọc dữ liệu từ Excel ----
    def doc_file(self):
        if not os.path.exists(self.filename):
            return
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active

        self.ds_nv = []  # làm mới danh sách
        for row in ws.iter_rows(min_row=2, values_only=True):
            stt, ma, ten, tuoi = row
            if ma and ten and tuoi:
                self.ds_nv.append(NhanVien(ma, ten, tuoi))
        wb.close()

    # ---- Ghi dữ liệu ra Excel ----
    def ghi_file(self):
        wb = Workbook()
